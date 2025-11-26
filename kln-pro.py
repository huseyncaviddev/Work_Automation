from pathlib import Path
import sys # Xəta idarəçiliyi üçün əlavə edildi

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Qalın kənar tərzi (THIN side style)
THIN = Side(border_style="thin", color="000000")

# Excel faylının konkret yaranacağı yer
# Yolu özünüzə uyğun dəyişdirə bilərsiniz.
OUTPUT_DIR = Path(r"C:\Users\husey\OneDrive\Desktop\Development\Work_Automation")
OUTPUT_PATH = OUTPUT_DIR / "SPP2-KLN-PRO-TRN-0164.xlsx"


def create_trn_0164_excel(output_path: Path = OUTPUT_PATH):
    """SPP2-KLN-PRO-TRN-0164 adlı Transmittal faylını yaradır."""
    
    # Kataloqu yaratmaq üçün cəhd, icazə problemini əvvəlcədən yoxlamaq üçün
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except PermissionError:
        print(f"❌ XƏTA: '{output_path.parent}' yoluna yazmaq üçün icazə yoxdur.")
        print("Zəhmət olmasa, yolu dəyişdirin və ya proqramı Administrator icazələri ilə işə salın.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ XƏTA: Kataloq yaratma zamanı naməlum xəta: {e}")
        sys.exit(1)


    wb = Workbook()
    ws = wb.active
    ws.title = "Transmittal"

    # ================== COLUMN WIDTHS ==================
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60

    # ================== TOP TITLES ==================
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4  # green line

    # Header 1
    ws.merge_cells("A1:F1")
    ws["A1"] = "SITALCHAY 2 PRODUCTION PLANT"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Header 2
    ws.merge_cells("A2:F2")
    ws["A2"] = "DOCUMENTATION TRANSMITTAL"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Green bar
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.fill = green_fill

    # ================== TRANS INFO BOX ==================
    for r in range(5, 8):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

    # Row 5
    ws["A5"] = "TRANSMITTAL NUMBER"
    ws["A5"].font = Font(bold=True)
    ws.merge_cells("B5:C5")
    ws["B5"] = "SPP2-KLN-PRO-TRN-0164"
    ws["D5"] = "DATE"
    ws["D5"].font = Font(bold=True)
    ws.merge_cells("E5:F5")
    ws["E5"] = "29-Jul-2025"
    
    # Row 6
    ws["A6"] = "PROJECT"
    ws["A6"].font = Font(bold=True)
    ws.merge_cells("B6:C6")
    ws["B6"] = "SPP2\nSITALCHAY 2 PRODUCTION PLANT"
    ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D6"] = "LOCATION"
    ws["D6"].font = Font(bold=True)
    ws.merge_cells("E6:F6")
    ws["E6"] = "SUMGAIT AZERBAIJAN"

    # Row 7 (Empty row inside box)
    ws.merge_cells("A7:F7")

    # ================== FROM / TO BLOKLARI ==================
    ws.row_dimensions[9].height = 18

    ws.merge_cells("A9:C9")
    ws["A9"] = "From:"
    ws["A9"].font = Font(bold=True)

    ws.merge_cells("D9:F9")
    ws["D9"] = "To:"
    ws["D9"].font = Font(bold=True)

    from_block = (
        '"KOLIN"  İNŞAAT SANAYI VE TICARET A.Ş\n'
        "Teoman Uludag\n"
        "Project Manager\n"
        "tuludag@kolin.com.tr"
    )

    to_block = (
        '"PROYAPI/PROKON" JV\n'
        "Mesut Sorgec\n"
        "Project Manager\n"
        "mesutsorgec@proyapimusavirlik.com"
    )

    # From block content
    ws.merge_cells("A10:C13")
    ws["A10"] = from_block
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")

    # To block content
    ws.merge_cells("D10:F13")
    ws["D10"] = to_block
    ws["D10"].alignment = Alignment(wrap_text=True, vertical="top")

    # From/To border-lər
    for r in range(9, 14):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

    # ================== DOCUMENT LIST HEADER ==================
    ws.merge_cells("A15:F15")
    ws["A15"] = "DOCUMENT LIST"
    ws["A15"].font = Font(bold=True)
    ws["A15"].alignment = Alignment(horizontal="center")

    header_row = 17
    headers = ["#", "Document Number", "Format", "Rev.", "Issue Code", "Document Title"]
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
        cell.fill = header_fill

    # ================== DOCUMENT ROWS ==================
    docs = [
        (1, "KLN-SPP2-ITP-CV-GN00-201", "PDF", "05", "IFA",
         "Inspection And Test Plan For Concrete And Insulation Works"),
        (2, "KLN-SPP2-MAR-AR-GN00-037", "PDF", "00", "IFA",
         "VesnaMetal Jacketing Starting U Profile"),
        (3, "KLN-SPP2-MAR-AR-GN00-038", "PDF", "00", "IFA",
         "Aluminium Verticale Profil 140/120/100/80"),
        (4, "KLN-SPP2-MAR-AR-GN00-039", "PDF", "00", "IFA",
         "Knauf Corner Profile"),
        (5, "KLN-SPP2-MAR-CV-GN00-065", "PDF", "00", "IFA",
         "Razor Wire"),
        (6, "KLN-SPP2-MAR-MC-GN00-072", "PDF", "00", "IFA",
         "Pipe Grooved Couplings"),
        (7, "KLN-SPP2-MAR-MC-GN00-073", "PDF", "00", "IFA",
         "Flexible Air Ducts"),
        # '*END*' sətrini sildim, çünki bu, fayl formatına uyğun deyil.
        # Əgər faylın sonunu göstərmək lazımdırsa, sətir nömrəsi olmadan boş sətir saxlanmalıdır.
    ]

    row = header_row + 1
    for no, doc_no, fmt, rev, issue, title in docs:
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=doc_no)
        ws.cell(row=row, column=3, value=fmt)
        ws.cell(row=row, column=4, value=rev)
        ws.cell(row=row, column=5, value=issue)
        ws.cell(row=row, column=6, value=title)

        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
            if col == 6:
                # Document Title üçün wrap və vertical top alignment
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row].height = 30 # Sətir hündürlüyünü artırırıq
            elif col == 2 or col == 5:
                # Document Number və Issue Code sola
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                # Digərləri mərkəzə
                c.alignment = Alignment(horizontal="center", vertical="center")

        row += 1
        
    # Ən son sətirə "END" qeydini əlavə etmək istəsəniz:
    ws.cell(row=row, column=1, value="*END*").alignment = Alignment(horizontal="center")
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
        
    row += 1 # Növbəti hissə üçün sətiri artırırıq


    # ================== ATTACHMENT & FOOTER ==================
    attach_row = row + 2
    ws.merge_cells(f"A{attach_row}:F{attach_row}")
    ws[f"A{attach_row}"] = "Attachment : ITP, MAR"

    footer_row_1 = attach_row + 3
    ws.merge_cells(f"A{footer_row_1}:F{footer_row_1}")
    ws[f"A{footer_row_1}"] = "VektorDS LLC | U.Hajibeyli str., 62, Baku, Azerbaijan. info@vektords.az"
    ws[f"A{footer_row_1}"].alignment = Alignment(wrap_text=True)

    footer_row_2 = footer_row_1 + 2
    ws.merge_cells(f"A{footer_row_2}:F{footer_row_2}")
    ws[f"A{footer_row_2}"] = (
        "Status Code: A = Accepted, AC = Accepted with Comments, CR = Commented-Resubmit, NA = Not Accepted; "
        "ADV = Advanced Copy, IFD = Issued For Design, IFI = Issued For Information, "
        "IFR = Issued For Review, IFA = Issued For Approval, IFC = Issued For Construction"
    )
    ws[f"A{footer_row_2}"].alignment = Alignment(wrap_text=True)

    # ================== FAYLIN SAXlanMASI ==================
    try:
        wb.save(output_path)
        print(f"✅ Uğurla yaradıldı: {output_path}")
    except PermissionError:
        print(f"❌ XƏTA: Faylı '{output_path}' yoluna yaza bilmədim.")
        print("Zəhmət olmasa, **faylın (SPP2-KLN-PRO-TRN-0164.xlsx) Microsoft Excel-də açıq olmadığını** yoxlayın və kodu yenidən işə salın.")
    except Exception as e:
        print(f"❌ XƏTA: Faylın saxlanması zamanı naməlum xəta: {e}")


def main():
    print("Working dir:", Path.cwd())
    print("Target file:", OUTPUT_PATH)
    create_trn_0164_excel()


if __name__ == "__main__":
    main()