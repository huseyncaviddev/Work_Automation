from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

# Excel hüceyrələrində istifadə olunacaq nazik (thin) sərhəd tərifi
THIN = Side(border_style="thin", color="000000")


def apply_border(ws, cell_range: str):
    """
    Verilən A1-stil intervala sərhəd (border) tətbiq edir.
    Məs: 'A1:Z10' kimi.
    """
    # ws[cell_range] → həmin aralıqdakı bütün sətirləri (tuple) qaytarır
    for row in ws[cell_range]:
        # Hər sətirdəki hər hüceyrəyə eyni border-i veririk
        for cell in row:
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def add_logos(ws, left_logo_path: Path, right_logo_path: Path):
    """
    Logoları header-in sol və sağ bloklarına yerləşdirir.
    - Sol logo: A2 hüceyrəsinin üzərinə anchor olunur (A2:E6 area).
    - Sağ logo: U2/Y6 tərəfə anchor olunur.
    """
    # Sol logo faylı varsa, şəkli sheet-ə əlavə et
    if left_logo_path.is_file():
        left_img = Image(str(left_logo_path))  # Şəkli yükləyirik
        left_img.width = 140                  # Şəkilin eni (px)
        left_img.height = 50                  # Şəkilin hündürlüyü (px)
        left_img.anchor = "A2"                # Şəkilin başlanğıc hüceyrəsi
        ws.add_image(left_img)                # Sheet-ə əlavə et

    # Sağ logo faylı varsa, onu da sheet-ə əlavə et
    if right_logo_path.is_file():
        right_img = Image(str(right_logo_path))
        right_img.width = 140
        right_img.height = 60
        right_img.anchor = "U2"               # Sağ tərəfdə başlanğıc hüceyrə
        ws.add_image(right_img)


def safe_save_workbook(wb: Workbook, output_path: Path) -> Path:
    """
    Faylı təhlükəsiz şəkildə saxlayır:
      - Əgər eyni adda fayl mövcuddursa, adın sonuna '_NEW' əlavə edir.
      - Folder yoxdursa, parent folderləri də yaradır.
    """
    output_path = Path(output_path)

    # Eyni adda fayl artıq mövcuddursa, yeni ad formalaşdır
    if output_path.exists():
        output_path = output_path.with_name(
            output_path.stem + "_NEW" + output_path.suffix
        )

    # Parent qovluqları (dirs) yoxdursa, yarat
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Workbook-u həmin path-ə save et
    wb.save(output_path)
    return output_path


def create_trn_excel(
    output_path: Path = Path("SPP2-KLN-PRO-TRN-0164_AUTO.xlsx"),
    trn_no: str = "SPP2-KLN-PRO-TRN-0164",
    date_str: str = "29-Jul-2025",
    left_logo: str = "vektords.png",
    right_logo: str = "proyapi_prokon.png",
):
    """
    TRN transmittal template-ini sıfırdan yaradan əsas funksiya.
    Parametrlərlə:
      - output_path  → çıxış faylının adı / yolu
      - trn_no       → transmittal nömrəsi
      - date_str     → tarix mətn formatında
      - left_logo    → sol logo fayl adı
      - right_logo   → sağ logo fayl adı
    """
    wb = Workbook()           # Yeni Excel workbook yaradırıq
    ws = wb.active            # Default olaraq açılan ilk sheet
    ws.title = "TRN Maker"    # Sheet-in adını dəyişirik

    # --- Ümumi default font və alignment ---
    base_font = Font(name="Calibri", size=9)
    # 1–59-cu sətirlər, 1–26-cı sütunlar (A–Z) üçün default stil veririk
    for row in range(1, 60):
        for col in range(1, 27):
            c = ws.cell(row=row, column=col)
            c.font = base_font
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # Sütun genişlikləri – sənin templatedəki layout-a yaxın
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["J"].width = 4
    ws.column_dimensions["K"].width = 4
    ws.column_dimensions["L"].width = 4
    ws.column_dimensions["M"].width = 4
    ws.column_dimensions["N"].width = 4
    ws.column_dimensions["O"].width = 4
    ws.column_dimensions["P"].width = 4
    ws.column_dimensions["Q"].width = 4
    ws.column_dimensions["R"].width = 4
    ws.column_dimensions["S"].width = 4
    ws.column_dimensions["T"].width = 4
    ws.column_dimensions["U"].width = 4
    ws.column_dimensions["V"].width = 4
    ws.column_dimensions["W"].width = 4
    ws.column_dimensions["X"].width = 4
    ws.column_dimensions["Y"].width = 4
    ws.column_dimensions["Z"].width = 4

    # Müəyyən sətirlərə xüsusi hündürlük veririk (layout matching)
    for r in range(11, 16):
        ws.row_dimensions[r].height = 16
    for r in range(19, 25):
        ws.row_dimensions[r].height = 17.5
    for r in range(21, 39):
        ws.row_dimensions[r].height = 18.5
    ws.row_dimensions[49].height = 13.5
    ws.row_dimensions[50].height = 13.5

    # === ÜST YAŞIL BAR / QUTU ===
    ws.merge_cells("A1:Z1")   # A1–Z1 aralığını bir hüceyrə kimi merge edirik
    apply_border(ws, "A1:Z1") # Sərhəd tətbiq edirik

    # Yaşıl dolgu (header bar rəngi – #91D050)
    fill_green = PatternFill(start_color="91D050", end_color="91D050", fill_type="solid")
    ws["A1"].fill = fill_green

    # === HEADER / TITLE AREA ===

    # Sol blok (logo, KOLIN və s. üçün boş ərazi) – A2:E6
    ws.merge_cells("A2:E6")
    apply_border(ws, "A2:E6")

    # Ortadakı başlıq – F2:U4
    ws.merge_cells("F2:U4")
    title = ws["F2"]
    title.value = "SITALCHAY 2 PRODUCTION PLANT\nDOCUMENTATION TRANSMITTAL"
    title.font = Font(name="Calibri", size=10, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(ws, "F2:U4")

    # Sağ blok (logo + page/rev) – V2:Z6
    ws.merge_cells("V2:Z6")
    apply_border(ws, "V2:Z6")

    # === TRANSMITTAL NUMBER sahəsi ===

    # “TRANSMITTAL NUMBER” label hissəsi
    ws.merge_cells("F5:M6")
    ws["F5"].value = "TRANSMITTAL  NUMBER:"
    ws["F5"].font = Font(name="Calibri", size=10, bold=True)
    ws["F5"].alignment = Alignment(horizontal="left", vertical="center", )

    # TRN nömrəsinin yazıldığı box
    ws.merge_cells("N5:U6")
    ws["N5"].value = trn_no
    ws["N5"].alignment = Alignment(horizontal="left", vertical="center")
    apply_border(ws, "F5:U6")

    # === DATE / PROJECT / LOCATION / PAGE / REV sətirləri ===

    # Sol – Date hissəsi (A7:E8)
    ws.merge_cells("A7:E8")
    ws["A7"].value = f"DATE: {date_str}"
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

    # PROJECT hissəsi (F7:M8)
    ws.merge_cells("F7:M8")
    ws["F7"].value = "PROJECT: SPP2 \nSITALCHAY 2 PRODUCTION PLANT "
    ws["F7"].alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

    # LOCATION hissəsi (N7:U8)
    ws.merge_cells("N7:U8")
    ws["N7"].value = "LOCATION: \nSUMGAIT AZERBAIJAN "
    ws["N7"].alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

    # Sağda Page info (V7:X8)
    ws.merge_cells("V7:X8")
    ws["V7"].value = "Page 1 of 1"
    ws["V7"].alignment = Alignment(horizontal="center", vertical="center")

    # Sağda Revision info (Y7:Z8)
    ws.merge_cells("Y7:Z8")
    ws["Y7"].value = "Rev.03"
    ws["Y7"].alignment = Alignment(horizontal="center", vertical="center")

    # A2–Z8 aralığına sərhəd əlavə edirik (bütün header blok)
    apply_border(ws, "A2:Z8")

############### === FROM / TO BLOKU ===###################

    # “From:” və “To:” başlıqları üçün sətir merge-ləri
    ws.merge_cells("A11:M11")
    ws.merge_cells("N11:Z11")
    ws["A11"].value = " From:"
    ws["N11"].value = " To:"

    # Şirkət adları
    ws.merge_cells("A12:M12")
    ws.merge_cells("N12:Z12")
    ws["A12"].value = ' “KOLIN”  İNŞAAT VE TICARET A.Ş'
    ws["N12"].value = ' “PROYAPI/PROKON” JV'

    # Adlar
    ws.merge_cells("A13:M13")
    ws.merge_cells("N13:Z13")
    ws["A13"].value = " Teoman Uludag"
    ws["N13"].value = " Mesut Sorgec"

    # Vəzifələr
    ws.merge_cells("A14:M14")
    ws.merge_cells("N14:Z14")
    ws["A14"].value = " Project Manager"
    ws["N14"].value = " Project Manager"

    # Email-lər
    ws.merge_cells("A15:M15")
    ws.merge_cells("N15:Z15")
    ws["A15"].value = " tuludag@kolin.com.tr"
    ws["N15"].value = " mesutsorgec@proyapimusavirlik.com"

    # FROM/TO blokuna border
    apply_border(ws, "A11:Z15")

    ############## === DOCUMENT LIST TITLE === ##############

    # “DOCUMENT LIST” başlığı üçün merge və format
    ws.merge_cells("J17:O17")
    ws["J17"].value = "DOCUMENT LIST"
    ws["J17"].font = Font(name="Calibri", size=10, bold=True)
    ws["J17"].alignment = Alignment(horizontal="center", vertical="center")

    # === DOCUMENT LIST TABLE HEADER ===

    # Header fon rəngi (boz)
    header_fill = PatternFill("solid", fgColor="FFE7E6E6")

    # Header cell-lərini strukturlaşdırmaq üçün merge-lər
    ws.merge_cells("A19:A20")   # Serial #
    ws.merge_cells("B19:G20")   # Document Number
    ws.merge_cells("H19:J20")   # Format
    ws.merge_cells("K19:L20")   # Rev.
    ws.merge_cells("M19:N20")   # Issue Code
    ws.merge_cells("O19:Z20")   # Document Title

    # Header text-lərini dictionary ilə yazırıq
    headers = {
        "A19": "#",
        "B19": "Document Number",
        "H19": "Format",
        "K19": "Rev.",
        "M19": "Issue\nCode",
        "O19": "Document Title",
    }

    # Hər header hüceyrəsinə text, font və background fill tətbiq edirik
    for cell_ref, text in headers.items():
        c = ws[cell_ref]
        c.value = text
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = header_fill

    apply_border(ws, "A19:Z20")

    ############### === DATA SƏTİRLƏRİ (21–38) === ###############

    # Hər data sətirindəki çox sütunlu sahələri merge edirik
    for row in range(21, 39):
        ws.merge_cells(f"B{row}:G{row}")  # Document Number area
        ws.merge_cells(f"H{row}:J{row}")  # Format area
        ws.merge_cells(f"K{row}:L{row}")  # Rev.
        ws.merge_cells(f"M{row}:N{row}")  # Issue Code
        ws.merge_cells(f"O{row}:Y{row}")  # Document Title

    # 21-ci sətrə nümunə data (sample row)
    ws["A21"].value = 1
    ws["A21"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B21"].value = "KLN-SPP2-ITP-CV-GN00-201"
    ws["H21"].value = "PDF"
    ws["K21"].value = "00"
    ws["M21"].value = "IFA"
    ws["O21"].value = "Inspection And Test Plan For Concrete And Insulation Works"

    # Mərkəzləndirilmiş alignment – format sahələri üçün
    for col in ("B", "H", "K", "M"):
        ws[f"{col}21"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Title soldan hizalı
    ws["O21"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Növbəti sətirlər üçün avtomatik sıra nömrəsi (formula ilə)
    for row in range(22, 39):
        ws[f"A{row}"].value = f"=A{row-1}+1"
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{row}"].font = Font(name="Calibri", size=8, bold=True)

    # END sətiri (39-cu sətir)
    ws["A39"].value = "=A38+1"
    ws["A39"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B39:G39")
    ws.merge_cells("H39:J39")
    ws.merge_cells("K39:L39")
    ws.merge_cells("M39:N39")
    ws.merge_cells("O39:Y39")
    ws["B39"].value = "*END*"
    ws["B39"].alignment = Alignment(horizontal="left", vertical="center")

    # Data table üçün sərhədlər
    apply_border(ws, "A21:Y39")

    # === FOOTER / ƏLAVƏ QEYDLƏR ===

    # Attachment sətiri
    ws.merge_cells("A41:Y41")
    ws["A41"].value = "Attachment: ITP, MAR"
    ws["A41"].alignment = Alignment(horizontal="left", vertical="center")
    apply_border(ws, "A41:Y41")

    # Status code izahı
    ws.merge_cells("A45:Y48")
    ws["A45"].value = (
        "Status Code: A = Accepted, AC = Accepted with Comments, CR = Commented-Resubmit, NA = Not Accepted\n"
        "ADV = Advanced Copy, IFD = Issued For Design, IFI = Issued For Information, IFR = Issued For Review, IFA = Issued For Approval\n"
        "IFC = Issued For Construction"
    )
    ws["A45"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(ws, "A45:Y48")

    # VektorDS footer mətni
    ws.merge_cells("A51:Y52")
    ws["A51"].value = (
        "VektorDS LLC | U.Hajibeyli str., 62, Baku, Azerbaijan. info@vektords.az\n"
        "This Document is VEKTORDS LLC property and cannot be used by others for any purpose without prior written consent."
    )
    ws["A51"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(ws, "A51:Y52")

    # === Logoları sheet-ə əlavə etmə ===

    # Cari script-in olduğu qovluq
    script_dir = Path(__file__).resolve().parent
    # Sol və sağ logo fayllarının tam yolu
    add_logos(ws, script_dir / left_logo, script_dir / right_logo)

    # Faylı təhlükəsiz şəkildə saxlayırıq
    saved = safe_save_workbook(wb, output_path)
    print(f"TRN Excel yaradıldı: {saved}")


if __name__ == "__main__":
    # Skript birbaşa çalışdırılarsa, default dəyərlərlə faylı yaradır
    create_trn_excel()
